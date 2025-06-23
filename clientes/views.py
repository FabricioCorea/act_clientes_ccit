from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User, Group
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from datetime import date, datetime
from django.utils.timezone import localdate
from clientes.models import *
from django.utils import timezone
import pytz
from django.db.models import Count
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from django.db.models import Max, F, Prefetch
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Exists, OuterRef, Q, Case, When, CharField, F, CharField, F, Q, Subquery, OuterRef, Value
from django.db.models import Subquery,  DateTimeField
from collections import defaultdict
from django.core.paginator import Paginator
from django.db import transaction
from itertools import chain
from django.db.models.functions import Coalesce, Greatest
from functools import reduce
from operator import or_




@login_required
def index(request):
    return render(request, 'inicio/inicio.html')

def paginar_queryset(request, queryset, param, por_pagina=10):
    paginator = Paginator(queryset, por_pagina)
    pagina = request.GET.get(param)
    return paginator.get_page(pagina)

@login_required
def clientes_pendientes(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get('q', '').strip()

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
 
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    clientes_pendientes_qs = Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_pendiente)
    if search_query:
        clientes_pendientes_qs = clientes_pendientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_pendientes = paginar_queryset(request, clientes_pendientes_qs, 'pendientes')

    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta") \
        .exclude(nombre__iexact="cerrado (por admin)")
    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()


    return render(request, 'clientes/clientes.html', {
        "clientes_pendientes": clientes_pendientes,
        "estado_reporte": estados_reporte,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_seguimiento(request):
    hoy = timezone.localdate()
    usuario = request.user

    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_query = request.GET.get("estado", "").strip().lower()

    # Estados considerados como seguimiento
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Subquery: última fecha del historial sin movimiento por cliente
    subquery_ultima_historial = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Query principal de clientes en seguimiento
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    ).annotate(
        ultima_fecha_hist=Subquery(subquery_ultima_historial, output_field=DateTimeField())
    )

    # Filtro por búsqueda
    if search_query:
        clientes_seguimiento_qs = clientes_seguimiento_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Filtro por estado
    if estado_query:
        clientes_seguimiento_qs = clientes_seguimiento_qs.filter(estado_actual__nombre__iexact=estado_query)

    # Ordenar por la última fecha del historial
    clientes_seguimiento_qs = clientes_seguimiento_qs.order_by('ultima_fecha_hist')

    # Paginación
    clientes_seguimiento = paginar_queryset(request, clientes_seguimiento_qs, 'seguimiento')

    # -------- Cálculo de clientes sin actualizar ----------
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    estados_reporte_filtro = EstadoReporte.objects.filter(
        id__in=Cliente.objects.filter(
            asignado_usuario=usuario,
            estado_actual__in=estados_seguimiento
        ).values_list('estado_actual', flat=True).distinct()
    )

    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta") \
        .exclude(nombre__iexact="cerrado (por admin)")

    # -------- Clientes actualizados hoy por el usuario ----------
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_qs.values('actualizado_por')[:1],
        primer_estado_nombre=primer_movimiento_qs.values('estado__nombre')[:1]
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_seguimiento": clientes_seguimiento,
        "estado_reporte": estados_reporte,
        "estado_reporte_filtro": estados_reporte_filtro,
        "estado_seguimiento": EstadoReporte.objects.filter(nombre__iexact="formulario enviado").first(),
        "view_type": "seguimiento",
        "search_query": search_query,
        "estado_query": estado_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado",
            movimientos__actualizado_por=usuario
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_sin_contestar(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Subquery para obtener la última fecha del historial sin movimiento
    subquery_ultima_historial = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Query principal con anotación y ordenamiento por fecha más antigua
    clientes_sin_contestar_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual=estado_no_contesto
    ).annotate(
        ultima_fecha_hist=Subquery(subquery_ultima_historial, output_field=DateTimeField())
    )

    if search_query:
        clientes_sin_contestar_qs = clientes_sin_contestar_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Ordenar por los más antiguos
    clientes_sin_contestar_qs = clientes_sin_contestar_qs.order_by('ultima_fecha_hist')

    # Paginación
    clientes_sin_contestar = paginar_queryset(request, clientes_sin_contestar_qs, 'nocontesto')

    # Contador sin actualizar
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    # Estados disponibles, excluyendo los no deseados
    estados_reporte = EstadoReporte.objects.filter(estado="activo") \
        .exclude(nombre__iexact="pendiente") \
        .exclude(nombre__iexact="no localizado") \
        .exclude(nombre__iexact="por localizar") \
        .exclude(nombre__iexact="formulario sin respuesta") \
        .exclude(nombre__iexact="cerrado (por admin)")
    
    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()
    
    return render(request, 'clientes/clientes.html', {
        "clientes_sin_contestar": clientes_sin_contestar,
        "estado_reporte": estados_reporte,
        "view_type": "nocontesto",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_sin_actualizar(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_filtrados_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).select_related('estado_actual')

    if search_query:
        clientes_filtrados_qs = clientes_filtrados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    # Subconsulta para obtener la fecha del último movimiento
    ultimo_movimiento_fecha = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Agrega anotación y ordena por la fecha más reciente
    clientes_filtrados_qs = clientes_filtrados_qs.annotate(
        ultima_fecha=Subquery(ultimo_movimiento_fecha)
    ).order_by('-ultima_fecha')

    clientes_resultado = paginar_queryset(request, clientes_filtrados_qs, 'sin_actualizar')

    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_completados": clientes_resultado,
        "view_type": "sin_actualizar",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(estado_actual__nombre__iexact="actualizado", movimientos__estado__nombre__iexact="actualizado", movimientos__actualizado_por=usuario).distinct().count(),
        "count_sin_actualizar": clientes_filtrados_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_actualizados(request):
    hoy = timezone.localdate() 
    usuario = request.user
    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    # Estados considerados como seguimiento: genera_movimiento=False y no están en excluidos
    estados_seguimiento = EstadoReporte.objects.filter(
        genera_movimiento=False
    ).exclude(nombre__iexact="pendiente").exclude(nombre__iexact="no contestó")

    # Clientes en seguimiento por esos estados
    clientes_seguimiento_qs = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__in=estados_seguimiento
    )
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    movimiento_actualizado_por_mi = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre__iexact="actualizado",
        actualizado_por=usuario
    )
    fecha_movimiento_actualizado = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre__iexact="actualizado"
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    clientes_actualizados_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado"
    ).annotate(
        fue_actualizado_por_mi=Exists(movimiento_actualizado_por_mi),
        fecha_actualizado=Subquery(fecha_movimiento_actualizado)
    ).filter(fue_actualizado_por_mi=True).select_related('estado_actual')

    if search_query:
        clientes_actualizados_qs = clientes_actualizados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    clientes_actualizados_qs = clientes_actualizados_qs.order_by('-fecha_actualizado')
    clientes_actualizados = paginar_queryset(request, clientes_actualizados_qs, 'actualizados')


    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    )

    # Clientes actualizados por el usuario hoy
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    # Clientes con movimiento hoy pero aún sin actualizar (cumplen lógica de sin_actualizar)
    movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=usuario)
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_usuario = Subquery(primer_movimiento_qs.values('actualizado_por')[:1])
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])
    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        tiene_movimiento_usuario=Exists(movimientos_usuario),
        primer_usuario_id=primer_movimiento_usuario,
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        (Q(tiene_movimiento_usuario=True) & ~Q(estado_actual__nombre__iexact="actualizado")) |
        (Q(estado_actual__nombre__iexact="actualizado") & Q(primer_usuario_id=usuario.id) & ~Q(primer_estado_nombre__iexact="actualizado"))
    ).distinct()

    return render(request, 'clientes/clientes.html', {
        "clientes_actualizados": clientes_actualizados,
        "estado_reporte": EstadoReporte.objects.filter(estado="activo").exclude(nombre__in=["pendiente", "no localizado", "por localizar", "formulario sin respuesta", "completado"]),
        "view_type": "actualizados",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(asignado_usuario=usuario, estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": clientes_actualizados_qs.count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@require_POST
def reportar_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    estado_id = request.POST.get("estado_reporte")
    nota_texto = request.POST.get("notas")
    estado_otro_nombre = request.POST.get("estado_otro")
    es_reenvio_directo = request.POST.get("accion_reenviar") == "1"

    if not cliente_id or not estado_id:
        messages.error(request, "Faltan datos del cliente o estado.")
        
        return redirect(request.META.get('HTTP_REFERER', '/'))


    cliente = get_object_or_404(Cliente, id=cliente_id)

    # Obtener o crear estado personalizado
    if estado_otro_nombre:
        estado = EstadoReporte.objects.create(
            nombre=estado_otro_nombre.strip(),
            creado_por=request.user,
            estado='activo',
            intentos=1,
            genera_movimiento=False
        )
    else:
        estado = get_object_or_404(EstadoReporte, id=estado_id)

    nombre_estado = estado.nombre.lower()

    # ------------------------ NO CONTESTÓ ------------------------
    if nombre_estado == "no contestó":
        cliente.sin_contestar += 1
        cliente.ultima_llamada_no_contesto = timezone.now()

        # Si aún no llega al límite, guardar historial sin movimiento
        if cliente.sin_contestar < estado.intentos:
            cliente.estado_actual = estado
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=False
            )

            messages.info(request, f"Intento de llamada {cliente.sin_contestar}/{estado.intentos}.")
            return redirect(request.META.get('HTTP_REFERER', '/'))


        # Si llega al límite, pasar a "por localizar"
        estado_por_localizar, _ = EstadoReporte.objects.get_or_create(
            nombre__iexact="por localizar",
            defaults={'nombre': 'Por localizar', 'creado_por': request.user, 'estado': 'activo'}
        )

        cliente.estado_actual = estado_por_localizar
        cliente.veces_contactado += 1
        cliente.sin_contestar = 0

        try:
            colector = User.objects.get(id=5)
            cliente.asignado_usuario = colector
        except User.DoesNotExist:
            pass

        cliente.save()

        # Guardar historial con genera_movimiento=True
        HistorialEstadoSinMovimiento.objects.create(
            cliente=cliente,
            estado=estado_por_localizar,
            actualizado_por=request.user,
            nota=nota_texto,
            genera_movimiento=True
        )

        movimiento = MovimientoEstado.objects.create(
            cliente=cliente,
            estado=estado_por_localizar,
            actualizado_por=request.user
        )

        if nota_texto:
            NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)

        messages.success(request, "Cliente enviado a colectores.")
        return redirect(request.META.get('HTTP_REFERER', '/'))


    # ------------------------ FORMULARIO ENVIADO ------------------------
    if nombre_estado == "formulario enviado":
        cliente.formulario_sin_contestar += 1
        cliente.ultimo_envio_formulario = timezone.now()

        # Si aún no llega al límite, guardar historial sin movimiento
        if cliente.formulario_sin_contestar < estado.intentos:
            cliente.estado_actual = estado
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=False
            )

            messages.info(request, f"Formulario enviado. Envío {cliente.formulario_sin_contestar}/{estado.intentos}.")
            return redirect(request.META.get('HTTP_REFERER', '/'))


        # Si llega al límite, pasar a "formulario sin respuesta"
        estado_sin_respuesta = EstadoReporte.objects.filter(nombre__iexact="formulario sin respuesta").first()
        if estado_sin_respuesta:
            cliente.estado_actual = estado_sin_respuesta
            cliente.formulario_sin_contestar = 0
            cliente.save()

            HistorialEstadoSinMovimiento.objects.create(
                cliente=cliente,
                estado=estado_sin_respuesta,
                actualizado_por=request.user,
                nota=nota_texto,
                genera_movimiento=True
            )

            movimiento = MovimientoEstado.objects.create(
                cliente=cliente,
                estado=estado_sin_respuesta,
                actualizado_por=request.user
            )

            if nota_texto:
                NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)

            messages.success(request, "Cliente actualizado con estado Formulario sin respuesta.")
            return redirect(request.META.get('HTTP_REFERER', '/'))


        messages.info(request, f"Formulario enviado. Envío {cliente.formulario_sin_contestar}/{estado.intentos}.")
        return redirect(request.META.get('HTTP_REFERER', '/'))


    # ------------------------ OTROS ESTADOS ------------------------
    cliente.veces_contactado += 1
    cliente.sin_contestar = 0
    cliente.formulario_sin_contestar = 0
    cliente.estado_actual = estado
    cliente.save()

    if estado.genera_movimiento:
        movimiento = MovimientoEstado.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=request.user
        )
        if nota_texto:
            NotaMovimiento.objects.create(movimiento=movimiento, texto=nota_texto)
        messages.success(request, "Cliente actualizado exitosamente.")
    else:
        HistorialEstadoSinMovimiento.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=request.user,
            nota=nota_texto,
            genera_movimiento=False
        )
        messages.info(request, "Cliente registrado en seguimiento. Este cliente aún no se actualiza.")

    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
@require_POST
def actualizar_estado_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    cliente = get_object_or_404(Cliente, id=cliente_id)

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
    if not estado_actualizado:
        messages.error(request, "Estado 'Actualizado' no encontrado.")
        return redirect(request.META.get('HTTP_REFERER', '/'))


    cliente.veces_contactado += 1
    cliente.sin_contestar = 0
    cliente.estado_actual = estado_actualizado
    cliente.save()

    MovimientoEstado.objects.create(
        cliente=cliente,
        estado=estado_actualizado,
        actualizado_por=request.user,
        fecha_hora=timezone.now()
    )

    messages.success(request, "Cliente actualizado exitosamente.")
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@require_POST
def crear_estado_reporte(request):
    nuevo_estado = request.POST.get("estado_otro", "").strip()

    if nuevo_estado:
        if EstadoReporte.objects.filter(nombre__iexact=nuevo_estado).exists():
            return JsonResponse({
                "success": False,
                "message": "Ya existe un estado con ese nombre.",
                "tags": "warning"
            })

        estado = EstadoReporte.objects.create(
            nombre=nuevo_estado,
            creado_por=request.user,
            estado="activo",
            intentos=1,
            genera_movimiento=False  # 👈🏽 Lo importante
        )

        return JsonResponse({
            "success": True,
            "id": estado.id,
            "nombre": estado.nombre,
            "message": "Estado creado exitosamente.",
            "tags": "success"
        })

    return JsonResponse({
        "success": False,
        "message": "Nombre de estado vacío.",
        "tags": "error"
    })

@login_required
def clientes_reportados(request):
    user = request.user
    grupos = user.groups.values_list('name', flat=True)
    search_query = request.GET.get('q', '').strip()
    estado_query = request.GET.get('estado', '').strip().lower()
    fecha_inicio = parse_date(request.GET.get('fecha_inicio', ''))
    fecha_fin = parse_date(request.GET.get('fecha_fin', ''))

    # Subqueries para la última fecha de actividad
    subquery_ultima_fecha_mov = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora').values('fecha_hora')[:1]
    subquery_ultima_fecha_hist = HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora').values('fecha_hora')[:1]

    clientes_reportados_query = Cliente.objects.prefetch_related(
        'movimientos__notas', 'historial_sin_movimiento', 'asignado_usuario', 'asignado_inicial', 'estado_actual'
    ).annotate(
        ultima_fecha_mov=Subquery(subquery_ultima_fecha_mov, output_field=DateTimeField()),
        ultima_fecha_hist=Subquery(subquery_ultima_fecha_hist, output_field=DateTimeField()),
        ultima_fecha=Greatest(
            Coalesce(Subquery(subquery_ultima_fecha_mov, output_field=DateTimeField()), Value(datetime(1900, 1, 1))),
            Coalesce(Subquery(subquery_ultima_fecha_hist, output_field=DateTimeField()), Value(datetime(1900, 1, 1)))
        )
    )

    # Filtro según rol
    if "super_admin" in grupos or "admin_group" in grupos:
        clientes_filtrados = clientes_reportados_query.filter(
            Q(Exists(MovimientoEstado.objects.filter(cliente=OuterRef('pk')))) |
            Q(Exists(HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk'))))
        )
    elif "estandar_group" in grupos or "colector_group" in grupos:
        clientes_filtrados = clientes_reportados_query.filter(
            Q(Exists(MovimientoEstado.objects.filter(cliente=OuterRef('pk'), actualizado_por=user))) |
            Q(Exists(HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk'), actualizado_por=user)))
        )
    else:
        messages.error(request, "Acceso no permitido.")
        return redirect("inicio")

    # Filtros de búsqueda y estado
    if search_query:
        clientes_filtrados = clientes_filtrados.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )
    if estado_query:
        clientes_filtrados = clientes_filtrados.filter(estado_actual__nombre__iexact=estado_query)
    if fecha_inicio and fecha_fin:
        clientes_filtrados = clientes_filtrados.filter(
            ultima_fecha__date__range=(fecha_inicio, fecha_fin)
        )

    # Orden y paginación
    clientes_filtrados = clientes_filtrados.order_by('-ultima_fecha')
    clientes_paginados = paginar_queryset(request, clientes_filtrados, 'reportados')

    # Procesamiento individual
    for cliente in clientes_paginados:
        movimientos_normales = [{"obj": m, "tipo": "con_movimiento"} for m in cliente.movimientos.all()]
        movimientos_historial = [{"obj": h, "tipo": "sin_movimiento"} for h in cliente.historial_sin_movimiento.all() if not h.genera_movimiento]
        todos_movimientos = sorted(movimientos_normales + movimientos_historial, key=lambda x: x["obj"].fecha_hora, reverse=True)

        movimientos_sin_admin = [m for m in todos_movimientos if getattr(m["obj"], "actualizado_por_admin", None) is None]

        cliente.todos_los_movimientos = todos_movimientos
        cliente.movimientos_sin_admin = movimientos_sin_admin

        # Último estado real y quién lo reportó
        if movimientos_sin_admin:
            cliente.ultimo_estado_real = movimientos_sin_admin[0]["obj"].estado
            cliente.reportado_por = movimientos_sin_admin[0]["obj"].actualizado_por
        else:
            cliente.ultimo_estado_real = cliente.estado_actual  # Fallback
            cliente.reportado_por = None

    # Estados disponibles (excluyendo pendiente)
    estados_disponibles = EstadoReporte.objects.filter(
        id__in=Cliente.objects.values_list('estado_actual', flat=True).distinct()
    ).exclude(nombre__iexact="pendiente")

    return render(request, 'clientes/clientes_reportados.html', {
        "clientes": clientes_paginados,
        "view_type": "reportados",
        "search_query": search_query,
        "estado_query": estado_query,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "estados_disponibles": estados_disponibles,
        "count_reportados": clientes_filtrados.count(),
    })

@login_required
def exportar_clientes_reportados_excel(request):
    if not request.user.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no autorizado.")
        return redirect("clientes_reportados")

    zona_honduras = pytz.timezone("America/Tegucigalpa")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Reportados"
    ws.append(["#", "Cliente", "Nombre", "Contacto", "Estado", "Usuario", "Fecha"])

    clientes = Cliente.objects.all().select_related("estado_actual", "asignado_usuario").order_by("numero_cliente")
    fila = 1

    for cliente in clientes:
        registros_cliente = []

        movimientos = [{"obj": m, "tipo": "con_movimiento"} for m in MovimientoEstado.objects.filter(cliente=cliente)]
        historiales = [{"obj": h, "tipo": "sin_movimiento"} for h in HistorialEstadoSinMovimiento.objects.filter(cliente=cliente, genera_movimiento=False)]

        todos = movimientos + historiales
        todos_ordenados = sorted(todos, key=lambda x: x["obj"].fecha_hora, reverse=True)

        buffer_repetidos = []
        estado_actual = None

        for idx, item in enumerate(todos_ordenados):
            estado_nombre = item["obj"].estado.nombre.strip().lower()
            tipo = item["tipo"]

            if tipo == "sin_movimiento":
                if not estado_actual or estado_nombre == estado_actual:
                    buffer_repetidos.append(item)
                    estado_actual = estado_nombre
                else:
                    # Se rompió la secuencia
                    if len(buffer_repetidos) >= 3:
                        # Mostrar intentos menos el más reciente
                        for i, intento in enumerate(reversed(buffer_repetidos[1:]), start=1):
                            registros_cliente.append({
                                "cliente": cliente,
                                "estado": f"{buffer_repetidos[0]['obj'].estado.nombre} (intento {i})",
                                "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                                "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                            })
                    else:
                        for i, intento in enumerate(reversed(buffer_repetidos), start=1):
                            registros_cliente.append({
                                "cliente": cliente,
                                "estado": f"{intento['obj'].estado.nombre} (intento {i})",
                                "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                                "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                            })
                    buffer_repetidos = [item]
                    estado_actual = estado_nombre
            else:
                # Procesar buffer antes del estado con movimiento
                if buffer_repetidos:
                    if len(buffer_repetidos) >= 3:
                        for i, intento in enumerate(reversed(buffer_repetidos[1:]), start=1):
                            registros_cliente.append({
                                "cliente": cliente,
                                "estado": f"{buffer_repetidos[0]['obj'].estado.nombre} (intento {i})",
                                "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                                "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                            })
                    else:
                        for i, intento in enumerate(reversed(buffer_repetidos), start=1):
                            registros_cliente.append({
                                "cliente": cliente,
                                "estado": f"{intento['obj'].estado.nombre} (intento {i})",
                                "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                                "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                            })
                    buffer_repetidos = []
                    estado_actual = None

                registros_cliente.append({
                    "cliente": cliente,
                    "estado": item["obj"].estado.nombre,
                    "usuario": item["obj"].actualizado_por.get_full_name() if item["obj"].actualizado_por else "Sin usuario",
                    "fecha": timezone.localtime(item["obj"].fecha_hora, zona_honduras),
                })

        # Procesar buffer al final
        if buffer_repetidos:
            if len(buffer_repetidos) >= 3:
                for i, intento in enumerate(reversed(buffer_repetidos[1:]), start=1):
                    registros_cliente.append({
                        "cliente": cliente,
                        "estado": f"{buffer_repetidos[0]['obj'].estado.nombre} (intento {i})",
                        "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                        "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                    })
            else:
                for i, intento in enumerate(reversed(buffer_repetidos), start=1):
                    registros_cliente.append({
                        "cliente": cliente,
                        "estado": f"{intento['obj'].estado.nombre} (intento {i})",
                        "usuario": intento["obj"].actualizado_por.get_full_name() if intento["obj"].actualizado_por else "Sin usuario",
                        "fecha": timezone.localtime(intento["obj"].fecha_hora, zona_honduras),
                    })

        # Ordenar registros del cliente por fecha descendente (ya lo están, pero aseguramos)
        registros_cliente = sorted(registros_cliente, key=lambda r: r["fecha"], reverse=True)

        # Agregar al Excel
        for reg in registros_cliente:
            ws.append([
                fila,
                reg["cliente"].numero_cliente,
                reg["cliente"].nombre_cliente,
                reg["cliente"].contacto_cliente,
                reg["estado"],
                reg["usuario"],
                reg["fecha"].strftime("%d/%m/%Y"),
            ])
            fila += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response["Content-Disposition"] = 'attachment; filename="clientes_tracking.xlsx"'
    wb.save(response)
    return response
@login_required
def dashboard_reportes(request):
    if not request.user.groups.filter(name__in=['super_admin', 'admin_group']).exists():
        if request.user.groups.filter(name="colector_group").exists():
            messages.error(request, "Acceso no permitido.")
            return redirect("clientes_colectores")
        elif request.user.groups.filter(name="estandar_group").exists():
            messages.error(request, "Acceso no permitido.")
            return redirect("clientes")

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario_id = request.GET.get('usuario_id')
    grupo = request.GET.get("grupo", "estandar")

    grupo_obj = {
        "estandar": Group.objects.get(name='estandar_group'),
        "colector": Group.objects.get(name='colector_group'),
    }
    grupo_activo = grupo_obj[grupo]
    usuarios = User.objects.filter(groups=grupo_activo).exclude(username="colector")

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
    estado_formulario = EstadoReporte.objects.filter(nombre__iexact="formulario enviado").first()
    estados_seguimiento = list(EstadoReporte.objects.filter(genera_movimiento=False).exclude(nombre__iexact="pendiente"))

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "usuario_id": usuario_id,
        "usuarios": usuarios,
        "grupo": grupo,
        "filtro_activo": bool(fecha_inicio or fecha_fin or usuario_id),
    }

    if grupo == "colector":
        usuarios_colectores = User.objects.filter(groups=grupo_activo)

        clientes_por_localizar_qs = Cliente.objects.filter(
            estado_actual__nombre__iexact="por localizar"
        )
        clientes_por_localizar_ids = set(clientes_por_localizar_qs.values_list("id", flat=True))

        clientes_contactados_ids = set()
        for cliente_id in clientes_por_localizar_ids:
            movs = MovimientoEstado.objects.filter(
                cliente_id=cliente_id,
                actualizado_por__in=usuarios_colectores,
                actualizado_por_admin__isnull=True
            ).order_by('fecha_hora')

            hist = HistorialEstadoSinMovimiento.objects.filter(
                cliente_id=cliente_id,
                actualizado_por__in=usuarios_colectores,
                actualizado_por_admin__isnull=True
            ).order_by('fecha_hora')

            registros = sorted(chain(movs, hist), key=lambda x: x.fecha_hora)

            posterior = False
            for r in registros:
                if r.estado.nombre.lower() == "por localizar":
                    posterior = True
                elif posterior:
                    clientes_contactados_ids.add(cliente_id)
                    break

        clientes_contactados_ids = list(clientes_contactados_ids)
        clientes_base = Cliente.objects.filter(id__in=clientes_por_localizar_ids)

        total_clientes = len(clientes_por_localizar_ids)
        clientes_contactados = len(clientes_contactados_ids)
        clientes_actualizados = Cliente.objects.filter(
            id__in=clientes_contactados_ids,
            estado_actual=estado_actualizado
        ).count()
        clientes_completados = Cliente.objects.filter(
            id__in=clientes_contactados_ids,
            estado_actual__genera_movimiento=True
        ).exclude(estado_actual=estado_actualizado).count()
        clientes_pendientes = total_clientes - clientes_contactados
        porcentaje_avance = round((clientes_contactados / total_clientes) * 100, 2) if total_clientes else 0

        reportes_por_usuario = defaultdict(int)
        actualizados_por_usuario = defaultdict(int)
        reportes_generales = defaultdict(int)

        for cliente_id in clientes_contactados_ids:
            registros = list(MovimientoEstado.objects.filter(
                cliente_id=cliente_id,
                actualizado_por__in=usuarios_colectores,
                actualizado_por_admin__isnull=True
            )) + list(HistorialEstadoSinMovimiento.objects.filter(
                cliente_id=cliente_id,
                actualizado_por__in=usuarios_colectores,
                actualizado_por_admin__isnull=True
            ))

            registros = sorted(registros, key=lambda x: x.fecha_hora)
            posterior = False
            for r in registros:
                if r.estado.nombre.lower() == "por localizar":
                    posterior = True
                elif posterior:
                    if r.actualizado_por:
                        nombre = r.actualizado_por.get_full_name()
                        reportes_por_usuario[nombre] += 1
                        if r.estado == estado_actualizado:
                            actualizados_por_usuario[nombre] += 1
                    reportes_generales[r.estado.nombre] += 1

        context.update({
            "clientes_totales_colector": total_clientes,
            "clientes_colector": clientes_contactados,
            "clientes_actualizados_colector": clientes_actualizados,
            "clientes_completados_colector": clientes_completados,
            "clientes_pendientes_colector": clientes_pendientes,
            "porcentaje_avance_colector": porcentaje_avance,
            "reportes_colector_por_estado": dict(reportes_generales),
            "reportes_colector_por_usuario": dict(reportes_por_usuario),
            "actualizados_colector_por_usuario": dict(actualizados_por_usuario),
        })

        return render(request, 'dashboard/dashboard.html', context)

    # === Grupo ESTÁNDAR ===

    clientes_base = Cliente.objects.filter(asignado_inicial__groups=grupo_activo)

    subquery_mov = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora').values('fecha_hora')[:1]
    subquery_hist = HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora').values('fecha_hora')[:1]

    mov_queryset = MovimientoEstado.objects.filter(actualizado_por_admin__isnull=True)
    hist_queryset = HistorialEstadoSinMovimiento.objects.filter(actualizado_por_admin__isnull=True)

    if usuario_id:
        mov_queryset = mov_queryset.filter(actualizado_por_id=usuario_id)
        hist_queryset = hist_queryset.filter(actualizado_por_id=usuario_id)

    clientes = clientes_base.annotate(
        ultima_fecha=Greatest(
            Coalesce(Subquery(subquery_mov, output_field=DateTimeField()), Value(datetime(1900, 1, 1))),
            Coalesce(Subquery(subquery_hist, output_field=DateTimeField()), Value(datetime(1900, 1, 1)))
        )
    ).only('id', 'estado_actual', 'asignado_inicial', 'asignado_usuario').prefetch_related(
        Prefetch('movimientos', queryset=mov_queryset.select_related('estado', 'actualizado_por'), to_attr='movimientos_validos'),
        Prefetch('historial_sin_movimiento', queryset=hist_queryset.select_related('estado', 'actualizado_por'), to_attr='historial_validos')
    )

    if fecha_inicio and fecha_fin:
        fecha_inicio_dt = parse_date(fecha_inicio)
        fecha_fin_dt = parse_date(fecha_fin)
        clientes = clientes.filter(ultima_fecha__date__range=(fecha_inicio_dt, fecha_fin_dt))

    reportes_generales = defaultdict(int)
    reportes_finalizados = defaultdict(int)
    reportes_seguimiento = defaultdict(int)
    reportes_por_usuario = defaultdict(int)
    actualizados_por_usuario = defaultdict(int)

    total_interacciones = interacciones_seguimiento = interacciones_formulario = interacciones_actualizados = interacciones_completados = 0
    clientes_contactados = clientes_actualizados = clientes_completados = clientes_en_seguimiento = 0

    for cliente in clientes:
        registros = sorted(
            chain(cliente.movimientos_validos, cliente.historial_validos),
            key=lambda x: x.fecha_hora,
            reverse=True
        )
        registros_validos = [r for r in registros if r.estado and r.actualizado_por]

        if registros_validos:
            ultimo_valido = registros_validos[0]
            clientes_contactados += 1
            total_interacciones += len(registros_validos)
            interacciones_formulario += sum(r.estado == estado_formulario for r in registros_validos)
            interacciones_actualizados += sum(r.estado == estado_actualizado for r in registros_validos)
            interacciones_completados += sum(r.estado.genera_movimiento and r.estado != estado_actualizado for r in registros_validos)
            interacciones_seguimiento += sum(r.estado in estados_seguimiento for r in registros_validos)

            if ultimo_valido.actualizado_por:
                nombre = ultimo_valido.actualizado_por.get_full_name()
                reportes_por_usuario[nombre] += 1
                if ultimo_valido.estado == estado_actualizado:
                    actualizados_por_usuario[nombre] += 1

        if cliente.estado_actual:
            reportes_generales[cliente.estado_actual.nombre] += 1
            if cliente.estado_actual == estado_actualizado:
                clientes_actualizados += 1
            elif cliente.estado_actual.genera_movimiento:
                clientes_completados += 1
            elif not cliente.estado_actual.genera_movimiento and cliente.estado_actual != estado_pendiente:
                clientes_en_seguimiento += 1

            if cliente.estado_actual.genera_movimiento:
                reportes_finalizados[cliente.estado_actual.nombre] += 1
            elif cliente.estado_actual.nombre.lower() != "pendiente":
                reportes_seguimiento[cliente.estado_actual.nombre] += 1

    total_clientes = Cliente.objects.count()
    clientes_asignados = clientes_base.distinct().count()
    clientes_pendientes = clientes_base.filter(estado_actual__nombre__iexact="pendiente").distinct().count()
    porcentaje_avance = round((clientes_actualizados / total_clientes) * 100, 2) if total_clientes else 0

    context.update({
        "clientes_totales": total_clientes,
        "clientes_estandar": clientes_contactados,
        "clientes_asignados_estandar": clientes_asignados,
        "clientes_actualizados_estandar": clientes_actualizados,
        "clientes_completados_estandar": clientes_completados,
        "clientes_en_seguimiento_estandar": clientes_en_seguimiento,
        "clientes_pendientes_estandar": clientes_pendientes,
        "avance_total_estandar": clientes_actualizados,
        "porcentaje_avance_estandar": porcentaje_avance,

        "total_interacciones_estandar": total_interacciones,
        "interacciones_estandar_seguimiento": interacciones_seguimiento,
        "interacciones_estandar_formulario": interacciones_formulario,
        "interacciones_estandar_actualizados": interacciones_actualizados,
        "interacciones_estandar_completados": interacciones_completados,

        "reportes_estandar_finalizados": dict(reportes_finalizados),
        "reportes_estandar_seguimiento": dict(reportes_seguimiento),
        "reportes_estandar_por_usuario": dict(reportes_por_usuario),
        "actualizados_estandar_por_usuario": dict(actualizados_por_usuario),
    })

    return render(request, 'dashboard/dashboard.html', context)
@login_required
def seguimiento_comparativa(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        return redirect("login")

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
    estado_formulario = EstadoReporte.objects.filter(nombre__iexact="formulario enviado").first()

    if not estado_actualizado or not estado_formulario:
        return render(request, "dashboard/seguimiento_comparativa.html", {})

    # Clientes actualmente en formulario enviado
    clientes_formulario_actual = Cliente.objects.filter(estado_actual=estado_formulario).distinct()
    clientes_formulario_actual_ids = set(clientes_formulario_actual.values_list('id', flat=True))

    # Total de veces que se registró el estado 'formulario enviado'
    total_envios_formulario = MovimientoEstado.objects.filter(estado=estado_formulario).count() + \
                              HistorialEstadoSinMovimiento.objects.filter(estado=estado_formulario).count()

    # Clientes que han tenido el estado 'formulario enviado'
    clientes_que_recibieron_formulario_qs = Cliente.objects.filter(
        Q(movimientos__estado=estado_formulario) | Q(historial_sin_movimiento__estado=estado_formulario)
    ).distinct()
    total_clientes_recibieron_formulario = clientes_que_recibieron_formulario_qs.count()

    # Clientes actualizados luego del formulario
    total_actualizados_con_formulario = clientes_que_recibieron_formulario_qs.filter(
        estado_actual=estado_actualizado
    ).count()

    # Clientes en estado actual "formulario enviado"
    total_sin_contestar_formulario = len(clientes_formulario_actual_ids)

    # Clientes que recibieron formulario y cambiaron de estado
    clientes_con_formulario_y_otro_estado = clientes_que_recibieron_formulario_qs.exclude(estado_actual=estado_formulario)
    total_formulario_y_otro_estado = clientes_con_formulario_y_otro_estado.count()

    # Frases de nota: no contestó
    frases_respuesta = [
        "responde", "responde.",
        "respondio", "respondio.",
        "respondió", "respondió.",
        "responder", "responder.",
        "contesta", "contesta.",
        "contestaron", "contestaron.",
        "no contesto", "no contesto.",
        "no contestó", "no contestó."
    ]
    texto_busqueda_respuesta = reduce(or_, [Q(nota__icontains=frase) for frase in frases_respuesta])
    historial_no_contesto = HistorialEstadoSinMovimiento.objects.filter(
        estado=estado_formulario,
        cliente__estado_actual=estado_formulario
    ).filter(texto_busqueda_respuesta).values_list("cliente_id", flat=True).distinct()
    ids_no_contesto = set(historial_no_contesto)

    # Frases de nota: teléfono fuera de línea, dañado, etc.
    frases_fuera_linea = [
        "fuera de linea", "fuera de linea.",
        "fuera de línea", "fuera de línea.",
        "fuera de servicio", "fuera de servicio.",
        "equivocado", "equivocado.",
        "disponible", "disponible.",
        "dañado", "dañado.",
        "dañada", "dañada.",
        "funciona", "funciona.",
        "mal estado", "mal estado.",
        "no es de la empresa", "no es de la empresa.",
        "no colabora", "no colabora."
    ]
    texto_busqueda_fuera_linea = reduce(or_, [Q(nota__icontains=frase) for frase in frases_fuera_linea])
    historial_fuera_linea = HistorialEstadoSinMovimiento.objects.filter(
        estado=estado_formulario,
        cliente__estado_actual=estado_formulario
    ).filter(texto_busqueda_fuera_linea).values_list("cliente_id", flat=True).distinct()
    ids_fuera_linea = set(historial_fuera_linea)

    # Frases de nota: solicitud por correo
    frases_envio_correo = [
        "via ", "vía ",
        "por correo", "por correo.",
        "verbal", "verbal.",
        "informacion", "información"
    ]
    texto_busqueda_envio_correo = reduce(or_, [Q(nota__icontains=frase) for frase in frases_envio_correo])
    historial_envio_correo = HistorialEstadoSinMovimiento.objects.filter(
        estado=estado_formulario,
        cliente__estado_actual=estado_formulario
    ).filter(texto_busqueda_envio_correo).values_list("cliente_id", flat=True).distinct()
    ids_por_correo = set(historial_envio_correo)

    # Clientes sin categoría registrada
    ids_sin_categoria = clientes_formulario_actual_ids - ids_no_contesto - ids_fuera_linea - ids_por_correo
    total_formularios_por_no_contestar = len(ids_no_contesto)
    total_formularios_fuera_linea = len(ids_fuera_linea)
    total_formularios_por_correo = len(ids_por_correo)
    total_formularios_sin_categoria = len(ids_sin_categoria)

    # Total de clientes en la base de datos
    total_clientes = Cliente.objects.count()

    # Porcentajes generales
    porcentaje_actualizados = round((total_actualizados_con_formulario / total_clientes_recibieron_formulario) * 100) if total_clientes_recibieron_formulario else 0
    porcentaje_sin_contestar = round((total_sin_contestar_formulario / total_clientes_recibieron_formulario) * 100) if total_clientes_recibieron_formulario else 0
    porcentaje_formulario_y_otro_estado = round((total_formulario_y_otro_estado / total_clientes_recibieron_formulario) * 100) if total_clientes_recibieron_formulario else 0

    # Porcentajes por categoría (solo para clientes en formulario enviado)
    porcentaje_no_contesto = round((total_formularios_por_no_contestar / total_sin_contestar_formulario) * 100) if total_sin_contestar_formulario else 0
    porcentaje_fuera_linea = round((total_formularios_fuera_linea / total_sin_contestar_formulario) * 100) if total_sin_contestar_formulario else 0
    porcentaje_por_correo = round((total_formularios_por_correo / total_sin_contestar_formulario) * 100) if total_sin_contestar_formulario else 0
    porcentaje_sin_categoria = round((total_formularios_sin_categoria / total_sin_contestar_formulario) * 100) if total_sin_contestar_formulario else 0

    # Porcentaje de clientes que recibieron formulario sobre el total
    porcentaje_recibieron_formulario = round((total_clientes_recibieron_formulario / total_clientes) * 100) if total_clientes else 0

    context = {
        "total_actualizados_con_formulario": total_actualizados_con_formulario,
        "total_sin_contestar_formulario": total_sin_contestar_formulario,
        "total_envios_formulario": total_envios_formulario,
        "clientes_que_recibieron_formulario": total_clientes_recibieron_formulario,
        "porcentaje_actualizados": porcentaje_actualizados,
        "porcentaje_sin_contestar": porcentaje_sin_contestar,
        "total_formularios_por_no_contestar": total_formularios_por_no_contestar,
        "total_formularios_fuera_linea": total_formularios_fuera_linea,
        "total_formularios_por_correo": total_formularios_por_correo,
        "total_formularios_sin_categoria": total_formularios_sin_categoria,
        "porcentaje_no_contesto": porcentaje_no_contesto,
        "porcentaje_fuera_linea": porcentaje_fuera_linea,
        "porcentaje_por_correo": porcentaje_por_correo,
        "porcentaje_sin_categoria": porcentaje_sin_categoria,
        "total_formulario_y_otro_estado": total_formulario_y_otro_estado,
        "porcentaje_formulario_y_otro_estado": porcentaje_formulario_y_otro_estado,
        "total_clientes": total_clientes,
        "porcentaje_recibieron_formulario": porcentaje_recibieron_formulario,
    }

    return render(request, "dashboard/seguimiento_comparativa.html", context)

@csrf_exempt
@login_required
def exportar_seguimiento_categoria(request):
    if request.method != "POST" or not request.user.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        return redirect("seguimiento_comparativa")

    categoria = request.POST.get("categoria")
    estado_formulario = EstadoReporte.objects.filter(nombre__iexact="formulario enviado").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    if not estado_formulario:
        return redirect("seguimiento_comparativa")

    zona_honduras = pytz.timezone("America/Tegucigalpa")
    clientes_ids = set()

    frases_dict = {
        "no_contesto": [
            "responde", "responde.",
            "respondio", "respondio.",
            "respondió", "respondió.",
            "responder", "responder.",
            "contesta", "contesta.",
            "contestaron", "contestaron.",
            "no contesto", "no contesto.",
            "no contestó", "no contestó."
        ],
        "fuera_linea": [
            "fuera de linea", "fuera de linea.",
            "fuera de línea", "fuera de línea.",
            "fuera de servicio", "fuera de servicio.",
            "equivocado", "equivocado.",
            "disponible", "disponible.",
            "dañado", "dañado.",
            "dañada", "dañada.",
            "funciona", "funciona.",
            "mal estado", "mal estado.",
            "no es de la empresa", "no es de la empresa.",
            "no colabora", "no colabora."
        ],
        "por_correo": [
            "via ", "vía ",
            "por correo", "por correo.",
            "verbal", "verbal."
            "informacion", "información"
        ]
    }

    if categoria == "recibieron":
        clientes_ids = set(
            Cliente.objects.filter(
                Q(movimientos__estado=estado_formulario) |
                Q(historial_sin_movimiento__estado=estado_formulario)
            ).distinct().values_list("id", flat=True)
        )

    elif categoria == "otro_estado":
        clientes_ids = set(
            Cliente.objects.filter(
                Q(movimientos__estado=estado_formulario) |
                Q(historial_sin_movimiento__estado=estado_formulario)
            ).exclude(estado_actual=estado_formulario).distinct().values_list("id", flat=True)
        )

    elif categoria == "actualizados":
        if estado_actualizado:
            clientes_ids = set(
                Cliente.objects.filter(
                    Q(movimientos__estado=estado_formulario) |
                    Q(historial_sin_movimiento__estado=estado_formulario),
                    estado_actual=estado_actualizado
                ).distinct().values_list("id", flat=True)
            )

    elif categoria == "sin_contestar":
        clientes_ids = set(
            Cliente.objects.filter(
                estado_actual=estado_formulario
            ).values_list("id", flat=True)
        )

    elif categoria in frases_dict:
        frases = frases_dict.get(categoria, [])
        query = reduce(Q.__or__, [Q(nota__icontains=f) for f in frases])
        clientes_ids = set(
            HistorialEstadoSinMovimiento.objects.filter(
                estado=estado_formulario,
                cliente__estado_actual=estado_formulario
            ).filter(query).values_list("cliente_id", flat=True).distinct()
        )

    elif categoria == "sin_categoria":
        todos = set(Cliente.objects.filter(estado_actual=estado_formulario).values_list("id", flat=True))

        frases_no_contesto = frases_dict["no_contesto"]
        frases_fuera_linea = frases_dict["fuera_linea"]
        frases_por_correo = frases_dict["por_correo"]

        q_no_contesto = reduce(Q.__or__, [Q(nota__icontains=f) for f in frases_no_contesto])
        q_fuera_linea = reduce(Q.__or__, [Q(nota__icontains=f) for f in frases_fuera_linea])
        q_por_correo = reduce(Q.__or__, [Q(nota__icontains=f) for f in frases_por_correo])

        ids_no_contesto = set(HistorialEstadoSinMovimiento.objects.filter(
            estado=estado_formulario,
            cliente__estado_actual=estado_formulario
        ).filter(q_no_contesto).values_list("cliente_id", flat=True).distinct())

        ids_fuera_linea = set(HistorialEstadoSinMovimiento.objects.filter(
            estado=estado_formulario,
            cliente__estado_actual=estado_formulario
        ).filter(q_fuera_linea).values_list("cliente_id", flat=True).distinct())

        ids_por_correo = set(HistorialEstadoSinMovimiento.objects.filter(
            estado=estado_formulario,
            cliente__estado_actual=estado_formulario
        ).filter(q_por_correo).values_list("cliente_id", flat=True).distinct())

        clientes_ids = todos - ids_no_contesto - ids_fuera_linea - ids_por_correo

    else:
        return redirect("seguimiento_comparativa")

    clientes = Cliente.objects.filter(id__in=clientes_ids).select_related("estado_actual", "asignado_usuario")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Exportados"
    ws.append(["#", "Cliente", "Nombre", "Contacto", "Estado", "Usuario", "Fecha", "Nota"])

    for i, cliente in enumerate(clientes, start=1):
        usuario = cliente.asignado_usuario.get_full_name() if cliente.asignado_usuario else "Sin asignar"
        fecha = ""
        nota_texto = ""

        mov = MovimientoEstado.objects.filter(
            cliente=cliente,
            estado=cliente.estado_actual
        ).order_by("-fecha_hora").first()

        hist = HistorialEstadoSinMovimiento.objects.filter(
            cliente=cliente,
            estado=cliente.estado_actual
        ).order_by("-fecha_hora").first()

        registro_mas_reciente = None
        if mov and hist:
            registro_mas_reciente = mov if mov.fecha_hora > hist.fecha_hora else hist
        elif mov:
            registro_mas_reciente = mov
        elif hist:
            registro_mas_reciente = hist

        if registro_mas_reciente:
            fecha = timezone.localtime(registro_mas_reciente.fecha_hora, zona_honduras).strftime("%d/%m/%Y %H:%M")

            if isinstance(registro_mas_reciente, MovimientoEstado):
                nota = NotaMovimiento.objects.filter(movimiento=registro_mas_reciente).order_by("fecha_creada").first()
                nota_texto = nota.texto if nota else ""
            elif isinstance(registro_mas_reciente, HistorialEstadoSinMovimiento):
                nota_texto = registro_mas_reciente.nota or ""

        ws.append([
            i,
            cliente.numero_cliente,
            cliente.nombre_cliente,
            cliente.contacto_cliente,
            cliente.estado_actual.nombre if cliente.estado_actual else "",
            usuario,
            fecha,
            nota_texto,
        ])

    nombres_personalizados = {
        "recibieron": "recibieron_correo",
        "otro_estado": "dieron_respuesta_a_correo",
        "actualizados": "actualizados_por_formulario",
        "sin_contestar": "sin_contestar_formulario",
        "sin_categoria": "sin_categoria_formulario_enviado",
        "no_contesto": "no_contesto_llamada_formulario_enviado",
        "fuera_linea": "telefonos_invalidos_formulario_enviado",
        "por_correo": "solicitaron_envio_correo"
    }

    nombre_archivo = f"clientes_{nombres_personalizados.get(categoria, categoria)}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response
@login_required
def clientes_sin_asignar_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()

    # Query base
    clientes_qs = Cliente.objects.filter(asignado_usuario__isnull=True)
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "sin_asignar")

    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs  = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.filter(
        movimientos__fecha_hora__date=hoy
    ).exclude(
        estado_actual__nombre__iexact="actualizado"
    ).distinct()

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()

    usuarios_con_clientes = User.objects.annotate(
        clientes_pendientes_count=Count('clientes_asignados', filter=Q(clientes_asignados__estado_actual=estado_pendiente))
    ).filter(clientes_pendientes_count__gt=0)


    usuarios_no_colectores = User.objects.exclude(groups__name="colector_group").exclude(username="rcoreas").filter(is_active=True)

    return render(request, "gestion/gestion.html", {
        "clientes_sin_asignar": clientes,
        "usuarios_no_colectores": usuarios_no_colectores,
        "view_type": "sin_asignar",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": Cliente.objects.exclude(
            estado_actual__nombre__iexact="actualizado"
        ).count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": clientes_qs.count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
        "usuarios_con_clientes": usuarios_con_clientes,
    })

@login_required
def clientes_actualizados_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Query base sin paginar (usada también para contadores)
    clientes_qs_base = Cliente.objects.filter(
        estado_actual=estado_actualizado,
        movimientos__estado=estado_actualizado
    ).distinct().prefetch_related('movimientos')

    # Filtro de búsqueda
    if search_query:
        clientes_qs_base = clientes_qs_base.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Aplicar paginación con prefijo 'actualizados'
    clientes = paginar_queryset(request, clientes_qs_base, "actualizados")

    # Contadores generales (no filtrados por usuario)
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, 'gestion/gestion.html', {
        "clientes_actualizados": clientes,
        "view_type": "actualizados",
        "search_query": search_query,
        "count_actualizados": clientes_qs_base.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_en_seguimiento_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")
    
    hoy = timezone.localdate()
    search_query = request.GET.get('q', '').strip()

    # Estados que definen a los clientes en seguimiento
    estados_seguimiento = EstadoReporte.objects.filter(genera_movimiento=False)\
        .exclude(nombre__iexact="pendiente")
        

    # Subquery para obtener la última fecha en HistorialEstadoSinMovimiento
    ultimo_historial_qs = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef("pk")
    ).order_by("-fecha_hora")

    # Query base para clientes en seguimiento con anotación de última fecha
    clientes_qs = Cliente.objects.filter(
        estado_actual__in=estados_seguimiento
    ).annotate(
        ultima_fecha_sin_movimiento=Subquery(
            ultimo_historial_qs.values("fecha_hora")[:1],
            output_field=DateTimeField()
        )
    )

    # Filtro por texto de búsqueda
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Paginación
    clientes = paginar_queryset(request, clientes_qs, 'seguimiento')

    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs = Cliente.objects.filter(
        estado_actual__in=estados_seguimiento
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, 'gestion/gestion.html', {
        "clientes_formulario_enviado": clientes,
        "view_type": "seguimiento",
        "search_query": search_query,
        "count_seguimiento": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_pendientes_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()

    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    # Query principal: todos los clientes con estado pendiente (sin filtrar por usuario)
    clientes_qs = Cliente.objects.filter(estado_actual=estado_pendiente)

    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "pendientes")

    # Estados de seguimiento (sin generar movimiento, excepto 'pendiente' y 'no contestó')
    estados_seguimiento = EstadoReporte.objects.filter(genera_movimiento=False)\
        .exclude(nombre__iexact="pendiente")
        

    clientes_seguimiento_qs = Cliente.objects.filter(estado_actual__in=estados_seguimiento)

    # Subqueries para detectar clientes actualizados pero cuyo primer estado no fue 'actualizado'
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    return render(request, "gestion/gestion.html", {
        "clientes_pendientes": clientes,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_qs.count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
    })

@login_required
def clientes_para_colectores_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")


    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()

    clientes_qs = Cliente.objects.filter(estado_actual=estado_por_localizar) if estado_por_localizar else Cliente.objects.none()

    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, "colectores")


    # Contadores generales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs  = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    usuarios_colectores = User.objects.filter(groups__name="colector_group", is_active=True)

    return render(request, "gestion/gestion.html", {
        "clientes_por_localizar": clientes,
        "usuarios_colectores": usuarios_colectores,
        "view_type": "colectores",
        "search_query": search_query,
        "count_colectores": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_todos": Cliente.objects.count(),
    })

login_required
def clientes_sin_actualizar_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    hoy = timezone.localdate()
    usuario = request.user

    if not usuario.groups.filter(name="estandar_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    search_query = request.GET.get("q", "").strip()

    # Estados definidos
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    # Obtener todos los estados que generan movimiento, excluyendo "actualizado"
    estados_con_movimiento = EstadoReporte.objects.filter(
        genera_movimiento=True
    ).exclude(nombre__iexact="actualizado")

    # Clientes cuyo estado actual es uno de los anteriores
    clientes_qs = Cliente.objects.filter(
        estado_actual__in=estados_con_movimiento
    ).select_related('estado_actual')

    # Subconsulta para obtener el primer estado del cliente
    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
    primer_estado_nombre = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

    # Subconsulta para obtener movimientos del cliente
    movimientos_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk'))

    # Anotaciones
    clientes_qs = clientes_qs.annotate(
        tiene_movimiento=Exists(movimientos_qs),
        primer_estado_nombre=primer_estado_nombre
    )

    # Filtro por búsqueda
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Anotaciones adicionales: fecha y usuario del último movimiento que coincide con el estado actual
    movimientos_compatibles = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk'),
        estado__nombre=OuterRef('estado_actual__nombre')
    ).order_by('-fecha_hora')

    clientes_qs = clientes_qs.annotate(
        ultima_fecha_movimiento=Subquery(
            movimientos_compatibles.values('fecha_hora')[:1],
            output_field=DateTimeField()
        ),
        usuario_movimiento=Subquery(
            movimientos_compatibles.values('actualizado_por__username')[:1]
        )
    ).order_by('-ultima_fecha_movimiento')

    clientes_resultado = paginar_queryset(request, clientes_qs, 'sin_actualizar')

    # Contadores
    clientes_seguimiento_qs = Cliente.objects.filter(estado_actual__in=estados_con_movimiento)
    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__actualizado_por=usuario,
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = clientes_qs.filter(
        ultima_fecha_movimiento__date=hoy
    )

    return render(request, 'clientes/clientes.html', {
        "clientes_completados": clientes_resultado,
        "view_type": "sin_actualizar_global",
        "search_query": search_query,
        "count_pendientes": Cliente.objects.filter(estado_actual__nombre__iexact="pendiente").count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual__nombre__iexact="no contestó").count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado",
            movimientos__actualizado_por=usuario
        ).distinct().count(),
        "count_sin_actualizar": clientes_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
    })

@login_required
def clientes_todos_view(request):
    usuario = request.user
    if not usuario.groups.filter(name__in=["super_admin", "admin_group"]).exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    hoy = timezone.localdate()
    search_query = request.GET.get("q", "").strip()
    estado_query = request.GET.get("estado", "").strip().lower()

    # Queryset base
    clientes_qs = Cliente.objects.select_related("estado_actual", "asignado_usuario")

    # Filtro por búsqueda
    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    # Filtro por estado
    if estado_query:
        clientes_qs = clientes_qs.filter(estado_actual__nombre__iexact=estado_query)

    # Paginación
    clientes = paginar_queryset(request, clientes_qs, "todos")

    # Cálculos adicionales
    estado_pendiente = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
    estado_no_contesto = EstadoReporte.objects.filter(nombre__iexact="no contestó").first()

    clientes_seguimiento_qs = Cliente.objects.filter(
        estado_actual__in=EstadoReporte.objects.filter(
            genera_movimiento=False
        ).exclude(nombre__iexact="pendiente")
    )

    primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef("pk")).order_by("fecha_hora")
    primer_movimiento_estado = Subquery(primer_movimiento_qs.values("estado__nombre")[:1])

    clientes_sin_actualizar_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    )

    actualizados_hoy_qs = Cliente.objects.filter(
        estado_actual__nombre__iexact="actualizado",
        movimientos__estado__nombre__iexact="actualizado",
        movimientos__fecha_hora__date=hoy
    ).distinct()

    clientes_sin_actualizar_hoy_qs = Cliente.objects.annotate(
        primer_estado_nombre=primer_movimiento_estado
    ).filter(
        Q(movimientos__fecha_hora__date=hoy),
        Q(estado_actual__nombre__iexact="actualizado") & ~Q(primer_estado_nombre__iexact="actualizado")
    ).distinct()

    # Estados disponibles para el filtro
    estados_disponibles = EstadoReporte.objects.filter(estado="activo").exclude(nombre="Completado")

    return render(request, "gestion/gestion.html", {
        "clientes_todos": clientes,
        "view_type": "todos",
        "search_query": search_query,
        "estado_query": estado_query,
        "estados_disponibles": estados_disponibles,
        "count_todos": clientes_qs.count(),
        "count_pendientes": Cliente.objects.filter(estado_actual=estado_pendiente).count(),
        "count_seguimiento": clientes_seguimiento_qs.count(),
        "count_no_contesto": Cliente.objects.filter(estado_actual=estado_no_contesto).count(),
        "count_actualizados": Cliente.objects.filter(
            estado_actual__nombre__iexact="actualizado",
            movimientos__estado__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_sin_actualizar": clientes_sin_actualizar_qs.count(),
        "count_actualizados_hoy": actualizados_hoy_qs.count() + clientes_sin_actualizar_hoy_qs.count(),
        "count_sin_asignar": Cliente.objects.filter(asignado_usuario__isnull=True).count(),
        "count_colectores": Cliente.objects.filter(estado_actual__nombre__iexact="por localizar").count(),
        "usuarios": User.objects.exclude(username__in=['rcoreas', 'colector']),
    })

@login_required
@require_POST
def asignar_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    usuario_id = request.POST.get("usuario_id")

    cliente = get_object_or_404(Cliente, id=cliente_id)
    usuario = get_object_or_404(User, id=usuario_id)

    # 👇 SOLO registrar la primera vez
    if cliente.asignado_inicial is None:
        cliente.asignado_inicial = usuario

    cliente.asignado_usuario = usuario
    cliente.save()

    messages.success(request, f"Cliente asignado exitosamente a {usuario.get_full_name()}.")
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@require_POST
def asignacion_por_cantidad(request):
    cantidad = int(request.POST.get('cantidad', 0))
    usuario_id = request.POST.get('usuario_id')

    if cantidad <= 0 or not usuario_id:
        messages.error(request, "Debe ingresar una cantidad válida y seleccionar un usuario.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    usuario = get_object_or_404(User, id=usuario_id)

    # No se permite asignar a usuarios del grupo colector
    if usuario.groups.filter(name="colector_group").exists():
        messages.error(request, "No se puede asignar clientes a usuarios del grupo 'colector'.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Obtener todos los clientes no asignados
    todos_no_asignados = Cliente.objects.filter(asignado_usuario__isnull=True).order_by('id')
    total_disponibles = todos_no_asignados.count()

    if total_disponibles == 0:
        messages.warning(request, "No hay clientes disponibles para asignar.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    if total_disponibles < cantidad:
        messages.warning(request, f"Solo hay {total_disponibles} cliente(s) sin asignar. Ajuste la cantidad.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Obtener solo los primeros `cantidad`
    clientes_a_asignar = todos_no_asignados[:cantidad]

    for cliente in clientes_a_asignar:
        # ✅ Solo asignar el inicial si aún no ha sido asignado antes
        if cliente.asignado_inicial is None:
            cliente.asignado_inicial = usuario

        cliente.asignado_usuario = usuario
        cliente.save()

    messages.success(request, f"{clientes_a_asignar.count()} clientes asignados a {usuario.get_full_name()}.")
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@require_POST
def reasignar_cliente_colector(request):
    cliente_id = request.POST.get("cliente_id")
    usuario_id = request.POST.get("usuario_id")

    cliente = get_object_or_404(Cliente, id=cliente_id)
    nuevo_usuario = get_object_or_404(User, id=usuario_id)

    # ✅ Solo se permite asignar a usuarios del grupo colector_group
    if not nuevo_usuario.groups.filter(name="colector_group").exists():
        messages.error(request, "Solo se puede asignar a usuarios del grupo 'colector_group'.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    cliente.asignado_usuario = nuevo_usuario
    cliente.save()

    messages.success(request, f"Cliente asignado exitosamente al colector {nuevo_usuario.get_full_name()}.")
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def clientes_colectores(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        if request.user.groups.filter(name="estandar_group").exists():
            return redirect("clientes")
        elif request.user.groups.filter(name__in=["super_admin", "admin_group"]).exists():
            return redirect("gestion")
        else:
            return redirect("login")

    # Estados necesarios
    estado_por_localizar = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Clientes con estado "por localizar" asignados al usuario
    clientes_por_localizar_asignados = Cliente.objects.none()
    clientes_por_localizar_filtrados = []

    if estado_por_localizar:
        clientes_por_localizar_asignados = Cliente.objects.filter(
            movimientos__estado=estado_por_localizar,
            asignado_usuario=request.user
        ).distinct().prefetch_related('movimientos')

        for cliente in clientes_por_localizar_asignados:
            movimientos_ordenados = sorted(cliente.movimientos.all(), key=lambda m: m.fecha_hora, reverse=True)
            if movimientos_ordenados and movimientos_ordenados[0].estado == estado_por_localizar:
                cliente.ultimo_movimiento = movimientos_ordenados[0]
                clientes_por_localizar_filtrados.append(cliente)

    # Movimientos con estado actualizado por el usuario
    movimientos_actualizados_por_usuario = MovimientoEstado.objects.none()
    if estado_actualizado:
        movimientos_actualizados_por_usuario = MovimientoEstado.objects.filter(
            estado=estado_actualizado,
            actualizado_por=request.user
        ).select_related('cliente', 'estado')

    # Estados visibles en el modal
    estados_incluir_modal = ["Actualizado", "Se negó", "No localizado", "Liquidada"]
    estados_visibles = EstadoReporte.objects.filter(nombre__in=estados_incluir_modal)

    # Clientes completados (reportados pero no actualizados)
    clientes_completados = Cliente.objects.filter(
        movimientos__actualizado_por=request.user
    ).distinct()

    if estado_actualizado:
        clientes_completados = clientes_completados.exclude(estado_actual=estado_actualizado)

    movimientos_prefetch = Prefetch(
        'movimientos',
        queryset=MovimientoEstado.objects.select_related('actualizado_por').order_by('-fecha_hora'),
        to_attr='movimientos_ordenados'
    )
    clientes_completados = clientes_completados.prefetch_related(movimientos_prefetch)

    for cliente in clientes_completados:
        cliente.ultimo_movimiento = cliente.movimientos_ordenados[0] if cliente.movimientos_ordenados else None

    return render(request, "clientes/clientes_colectores.html", {
        "clientes_por_localizar_filtrados": clientes_por_localizar_filtrados,
        "movimientos_actualizados_por_usuario": movimientos_actualizados_por_usuario,
        "clientes_completados": clientes_completados,
        "estado_reporte": estados_visibles
    })


@login_required
def clientes_colectores_pendientes(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()

    # Último estado en MovimientoEstado
    ult_estado_mov_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_mov_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Último estado en HistorialEstadoSinMovimiento
    ult_estado_hist_subquery = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_hist_subquery = HistorialEstadoSinMovimiento.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Base queryset
    clientes_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ult_estado_mov=Subquery(ult_estado_mov_subquery),
        ult_fecha_mov=Subquery(ult_fecha_mov_subquery),
        ult_estado_hist=Subquery(ult_estado_hist_subquery),
        ult_fecha_hist=Subquery(ult_fecha_hist_subquery),
    ).annotate(
        ultimo_estado=Case(
            When(Q(ult_fecha_mov__isnull=False) & Q(ult_fecha_hist__isnull=True), then=F('ult_estado_mov')),
            When(Q(ult_fecha_hist__isnull=False) & Q(ult_fecha_mov__isnull=True), then=F('ult_estado_hist')),
            When(Q(ult_fecha_mov__gte=F('ult_fecha_hist')), then=F('ult_estado_mov')),
            When(Q(ult_fecha_hist__gt=F('ult_fecha_mov')), then=F('ult_estado_hist')),
            default=Value(''),
            output_field=CharField()
        ),
        ultima_fecha=Case(
            When(Q(ult_fecha_mov__isnull=False) & Q(ult_fecha_hist__isnull=True), then=F('ult_fecha_mov')),
            When(Q(ult_fecha_hist__isnull=False) & Q(ult_fecha_mov__isnull=True), then=F('ult_fecha_hist')),
            When(Q(ult_fecha_mov__gte=F('ult_fecha_hist')), then=F('ult_fecha_mov')),
            When(Q(ult_fecha_hist__gt=F('ult_fecha_mov')), then=F('ult_fecha_hist')),
            default=None,
        )
    ).filter(
        Q(ultimo_estado__iexact="por localizar") | Q(ultimo_estado__iexact="pendiente")
    ).order_by(
        F('ultima_fecha').desc(nulls_last=True)
    )

    if search_query:
        clientes_qs = clientes_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes = paginar_queryset(request, clientes_qs, 'pendientes')

    return render(request, "clientes/clientes_colectores.html", {
        "clientes": clientes,
        "view_type": "pendientes",
        "search_query": search_query,
        "count_pendientes": clientes_qs.count(),
        "count_completados": Cliente.objects.filter(
            movimientos__actualizado_por=usuario
        ).exclude(
            estado_actual__nombre__iexact="actualizado"
        ).distinct().count(),
        "count_actualizados": MovimientoEstado.objects.filter(
            estado__nombre__iexact="actualizado", actualizado_por=usuario
        ).count(),
        "estado_reporte": EstadoReporte.objects.filter(
            nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"]
        )
    })

@login_required
def clientes_colectores_completados(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    # Subqueries para última fecha y último estado
    ult_estado_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    # Clientes completados (reportados pero no actualizados)
    clientes_completados_qs = Cliente.objects.filter(
        movimientos__actualizado_por=usuario
    ).exclude(
        estado_actual=estado_actualizado
    ).annotate(
        ultima_fecha=Subquery(ult_fecha_subquery),
        ultimo_estado=Subquery(ult_estado_subquery)
    ).order_by('-ultima_fecha').distinct()

    # Prefetch de movimientos ordenados
    movimientos_prefetch = Prefetch(
        'movimientos',
        queryset=MovimientoEstado.objects.select_related('estado').order_by('-fecha_hora'),
        to_attr='movimientos_ordenados'
    )
    clientes_completados_qs = clientes_completados_qs.prefetch_related(movimientos_prefetch)

    if search_query:
        clientes_completados_qs = clientes_completados_qs.filter(
            Q(nombre_cliente__icontains=search_query) |
            Q(numero_cliente__icontains=search_query) |
            Q(contacto_cliente__icontains=search_query)
        )

    clientes_completados = paginar_queryset(request, clientes_completados_qs, 'completados')

    # Asignar último movimiento para visualización
    for cliente in clientes_completados:
        cliente.ultimo_movimiento = cliente.movimientos_ordenados[0] if cliente.movimientos_ordenados else None

    # Contadores coherentes
    clientes_pendientes_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ultimo_estado=Subquery(ult_estado_subquery)
    ).filter(
        ultimo_estado__iexact="por localizar"
    )

    return render(request, "clientes/clientes_colectores.html", {
        "clientes": clientes_completados,
        "view_type": "completados",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_completados": clientes_completados_qs.count(),
        "count_actualizados": MovimientoEstado.objects.filter(estado=estado_actualizado, actualizado_por=usuario).count(),
        "estado_reporte": EstadoReporte.objects.filter(nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"])
    })

@login_required
def clientes_colectores_actualizados(request):
    if not request.user.groups.filter(name="colector_group").exists():
        messages.error(request, "Acceso no permitido.")
        return redirect("login")

    usuario = request.user
    search_query = request.GET.get("q", "").strip()

    estado_actualizado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()

    movimientos_actualizados_qs = MovimientoEstado.objects.filter(
        estado=estado_actualizado,
        actualizado_por=usuario
    ).select_related("cliente")

    if search_query:
        movimientos_actualizados_qs = movimientos_actualizados_qs.filter(
            Q(cliente__nombre_cliente__icontains=search_query) |
            Q(cliente__numero_cliente__icontains=search_query) |
            Q(cliente__contacto_cliente__icontains=search_query)
        )

    movimientos_actualizados = paginar_queryset(request, movimientos_actualizados_qs.order_by('-fecha_hora'), 'actualizados')

    # Subqueries para contar pendientes
    ult_estado_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('estado__nombre')[:1]

    ult_fecha_subquery = MovimientoEstado.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha_hora').values('fecha_hora')[:1]

    clientes_pendientes_qs = Cliente.objects.filter(
        asignado_usuario=usuario
    ).annotate(
        ultimo_estado=Subquery(ult_estado_subquery),
        ultima_fecha=Subquery(ult_fecha_subquery)
    ).filter(
        ultimo_estado__iexact="por localizar"
    )

    clientes_completados_qs = Cliente.objects.filter(
        movimientos__actualizado_por=usuario
    ).exclude(estado_actual=estado_actualizado).distinct()

    return render(request, "clientes/clientes_colectores.html", {
        "movimientos_actualizados": movimientos_actualizados,
        "view_type": "actualizados",
        "search_query": search_query,
        "count_pendientes": clientes_pendientes_qs.count(),
        "count_completados": clientes_completados_qs.count(),
        "count_actualizados": movimientos_actualizados_qs.count(),
        "estado_reporte": EstadoReporte.objects.filter(nombre__in=["Actualizado", "Se negó", "No localizado", "Liquidada"])
    })

@login_required
@require_POST
def importar_clientes(request):
    archivo = request.FILES.get("archivo_excel")
    if not archivo:
        messages.error(request, "Debe subir un archivo Excel.")
        return redirect("gestion")

    try:
        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active

        filas_insertadas = 0
        filas_actualizadas = 0
        errores = []

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            try:
                numero, nombre, direccion, contacto, contacto_cargo, telefono, telefono_dos, correo = fila[:8]
            except ValueError:
                errores.append(f"Fila {i}: Número incorrecto de columnas.")
                continue

            if not (numero and nombre):
                errores.append(f"Fila {i}: número o nombre faltante.")
                continue

            cliente = Cliente.objects.filter(numero_cliente=numero).first()
            estado = EstadoReporte.objects.filter(id=1).first()  # Estado "Pendiente"

            if not estado:
                errores.append(f"Fila {i}: No se encontró el estado con ID 1 (Pendiente).")
                continue

            if cliente:
                cliente.nombre_cliente = nombre
                cliente.contacto_cliente = contacto
                cliente.contacto_cargo = contacto_cargo or None
                cliente.telefono_cliente = str(telefono)
                cliente.telefono_dos = str(telefono_dos) if telefono_dos else None
                cliente.correo = correo or None
                cliente.direccion = direccion or None
                cliente.estado_actual = estado
                cliente.save()
                filas_actualizadas += 1
            else:
                Cliente.objects.create(
                    numero_cliente=numero,
                    nombre_cliente=nombre,
                    contacto_cliente=contacto,
                    contacto_cargo=contacto_cargo or None,
                    telefono_cliente=str(telefono),
                    telefono_dos=str(telefono_dos) if telefono_dos else None,
                    correo=correo or None,
                    direccion=direccion or None,
                    estado_actual=estado
                )
                filas_insertadas += 1

        msg = f"Importación completada: {filas_insertadas} insertado(s), {filas_actualizadas} actualizado(s)."
        if errores:
            msg += f" {len(errores)} error(es) encontrados. Revisa el archivo."

        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f"Error al procesar el archivo: {str(e)}")

    return redirect("gestion")

@login_required
@require_POST
def exportar_clientes(request):
    filtro = request.POST.get("filtro_exportacion")
    fecha_inicio = request.POST.get("fecha_inicio")
    fecha_fin = request.POST.get("fecha_fin")

    queryset = Cliente.objects.all().select_related("estado_actual", "asignado_usuario")
    zona_honduras = pytz.timezone("America/Tegucigalpa")

    if filtro == "actualizados":
        estado = EstadoReporte.objects.filter(nombre__iexact="actualizado").first()
        if estado:
            queryset = queryset.filter(movimientos__estado=estado).distinct()

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)

    elif filtro == "seguimiento":
        estados = EstadoReporte.objects.filter(genera_movimiento=False).exclude(nombre__iexact="pendiente")
        queryset = queryset.filter(estado_actual__in=estados)

        ultimo_historial_qs = HistorialEstadoSinMovimiento.objects.filter(
            cliente=OuterRef("pk")
        ).order_by("-fecha_hora")

        queryset = queryset.annotate(
            ultima_fecha_sin_movimiento=Subquery(
                ultimo_historial_qs.values("fecha_hora")[:1],
                output_field=DateTimeField()
            )
        )

        if fecha_inicio:
            queryset = queryset.filter(
                Q(ultimo_envio_formulario__date__gte=fecha_inicio) |
                Q(ultima_fecha_sin_movimiento__date__gte=fecha_inicio)
            )
        if fecha_fin:
            queryset = queryset.filter(
                Q(ultimo_envio_formulario__date__lte=fecha_fin) |
                Q(ultima_fecha_sin_movimiento__date__lte=fecha_fin)
            )

    elif filtro == "colectores":
        estado = EstadoReporte.objects.filter(nombre__iexact="por localizar").first()
        if estado:
            queryset = queryset.filter(estado_actual=estado)

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)

    elif filtro == "pendientes":
        estado = EstadoReporte.objects.filter(nombre__iexact="pendiente").first()
        if estado:
            queryset = queryset.filter(estado_actual=estado)

        if fecha_inicio:
            queryset = queryset.filter(movimientos__fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(movimientos__fecha_hora__date__lte=fecha_fin)

    elif filtro == "completados":
        estados_con_movimiento = EstadoReporte.objects.filter(
            genera_movimiento=True
        ).exclude(nombre__iexact="actualizado")

        movimientos_usuario = MovimientoEstado.objects.filter(cliente=OuterRef('pk'))
        primer_movimiento_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('fecha_hora')
        primer_estado_nombre = Subquery(primer_movimiento_qs.values('estado__nombre')[:1])

        movimientos_compatibles = MovimientoEstado.objects.filter(
            cliente=OuterRef('pk'),
            estado__nombre=OuterRef('estado_actual__nombre')
        ).order_by('-fecha_hora')

        queryset = Cliente.objects.filter(
            estado_actual__in=estados_con_movimiento
        ).annotate(
            tiene_movimiento=Exists(movimientos_usuario),
            primer_estado_nombre=primer_estado_nombre,
            ultima_fecha_movimiento=Subquery(
                movimientos_compatibles.values('fecha_hora')[:1],
                output_field=DateTimeField()
            ),
            usuario_movimiento=Subquery(
                movimientos_compatibles.values('actualizado_por__username')[:1]
            )
        ).order_by('-ultima_fecha_movimiento')

        if fecha_inicio:
            queryset = queryset.filter(ultima_fecha_movimiento__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(ultima_fecha_movimiento__date__lte=fecha_fin)

    elif filtro == "todos":
        # Anotar fecha del último MovimientoEstado
        ultimo_mov_qs = MovimientoEstado.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora')
        fecha_movimiento = Subquery(ultimo_mov_qs.values('fecha_hora')[:1], output_field=DateTimeField())

        # Anotar fecha del último HistorialEstadoSinMovimiento
        ultimo_hist_qs = HistorialEstadoSinMovimiento.objects.filter(cliente=OuterRef('pk')).order_by('-fecha_hora')
        fecha_historial = Subquery(ultimo_hist_qs.values('fecha_hora')[:1], output_field=DateTimeField())

        queryset = queryset.annotate(
            fecha_ultimo_movimiento=fecha_movimiento,
            fecha_ultimo_historial=fecha_historial
        )

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes Exportados"
    ws.append(["#", "Cliente", "Nombre", "Contacto", "Estado", "Usuario", "Fecha"])

    for i, cliente in enumerate(queryset, start=1):
        fecha = ""
        usuario = cliente.asignado_usuario.get_full_name() if cliente.asignado_usuario else "Sin asignar"

        if filtro == "seguimiento":
            if cliente.ultimo_envio_formulario:
                fecha = timezone.localtime(cliente.ultimo_envio_formulario, zona_honduras).strftime("%d/%m/%Y %H:%M")
            elif cliente.ultima_fecha_sin_movimiento:
                fecha = timezone.localtime(cliente.ultima_fecha_sin_movimiento, zona_honduras).strftime("%d/%m/%Y %H:%M")

        elif filtro == "completados":
            if cliente.ultima_fecha_movimiento:
                fecha = timezone.localtime(cliente.ultima_fecha_movimiento, zona_honduras).strftime("%d/%m/%Y %H:%M")
            usuario = cliente.usuario_movimiento if cliente.usuario_movimiento else "Sin asignar"

        elif filtro != "todos":
            ultimo_mov = cliente.movimientos.order_by("-fecha_hora").first()
            if ultimo_mov:
                fecha = timezone.localtime(ultimo_mov.fecha_hora, zona_honduras).strftime("%d/%m/%Y %H:%M")

        else:  # filtro == "todos"
            if cliente.estado_actual and cliente.estado_actual.nombre.lower() == "pendiente":
                fecha = ""
            else:
                fechas = [f for f in [cliente.fecha_ultimo_movimiento, cliente.fecha_ultimo_historial] if f]
                if fechas:
                    fecha_mas_reciente = max(fechas)
                    fecha = timezone.localtime(fecha_mas_reciente, zona_honduras).strftime("%d/%m/%Y %H:%M")

        ws.append([
            i,
            cliente.numero_cliente,
            cliente.nombre_cliente,
            cliente.contacto_cliente,
            cliente.estado_actual.nombre if cliente.estado_actual else "",
            usuario,
            fecha,
        ])

    # Nombre del archivo
    nombre_base = f"Clientes_{filtro}"
    if fecha_inicio and fecha_fin:
        nombre_archivo = f"{nombre_base}_{fecha_inicio}_a_{fecha_fin}.xlsx"
    elif fecha_inicio:
        nombre_archivo = f"{nombre_base}_desde_{fecha_inicio}.xlsx"
    elif fecha_fin:
        nombre_archivo = f"{nombre_base}_hasta_{fecha_fin}.xlsx"
    else:
        nombre_archivo = f"{nombre_base}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    wb.save(response)
    return response

@login_required
@require_POST
@transaction.atomic
def editar_cliente(request):
    cliente_id = request.POST.get("cliente_id")
    estado_nombre = request.POST.get("estado")
    motivo = request.POST.get("motivo")
    usuario_id = request.POST.get("usuario_id")  # para asignar
    acreditar_id = request.POST.get("acreditar_id")  # para acreditar

    cliente = get_object_or_404(Cliente, id=cliente_id)
    estado = get_object_or_404(EstadoReporte, nombre__iexact=estado_nombre)
    estado_anterior = cliente.estado_actual

    if estado_nombre == 'pendiente':
        # ✅ Obtener último usuario antes de borrar
        ultimo_movimiento = MovimientoEstado.objects.filter(cliente=cliente).order_by('-fecha_hora').first()
        ultimo_historial = HistorialEstadoSinMovimiento.objects.filter(cliente=cliente).order_by('-fecha_hora').first()

        ultimo_usuario = None
        if ultimo_movimiento and ultimo_movimiento.actualizado_por:
            ultimo_usuario = ultimo_movimiento.actualizado_por
        elif ultimo_historial and ultimo_historial.actualizado_por:
            ultimo_usuario = ultimo_historial.actualizado_por
        else:
            ultimo_usuario = request.user  # fallback

        # Borrar movimientos previos
        NotaMovimiento.objects.filter(movimiento__cliente=cliente).delete()
        MovimientoEstado.objects.filter(cliente=cliente).delete()
        HistorialEstadoSinMovimiento.objects.filter(cliente=cliente).delete()

        # Asignar usuario si se indica
        if usuario_id:
            asignado = get_object_or_404(User, id=usuario_id)
            cliente.asignado_usuario = asignado

        cliente.estado_actual = estado
        cliente.veces_contactado = 0
        cliente.sin_contestar = 0
        cliente.formulario_sin_contestar = 0
        cliente.save()

        # Registrar como historial sin movimiento
        HistorialEstadoSinMovimiento.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=ultimo_usuario,
            actualizado_por_admin=request.user if request.user.is_staff else None,
            nota=motivo or "Cambio a pendiente",
            genera_movimiento=False
        )

    elif estado_nombre in ['actualizado', 'Cerrado (por admin)']:

        if acreditar_id:
            acreditar_usuario = get_object_or_404(User, id=acreditar_id)
        else:
            acreditar_usuario = request.user

        movimiento = MovimientoEstado.objects.create(
            cliente=cliente,
            estado=estado,
            actualizado_por=acreditar_usuario,
            actualizado_por_admin=request.user
        )

        if motivo:
            NotaMovimiento.objects.create(movimiento=movimiento, texto=motivo)

        cliente.estado_actual = estado
        cliente.veces_contactado += 1
        cliente.sin_contestar = 0
        cliente.formulario_sin_contestar = 0
        cliente.save()

    else:
        cliente.estado_actual = estado
        cliente.veces_contactado += 1
        cliente.save()

    # Registrar motivo de cambio
    MotivoCambioEstado.objects.create(
        cliente=cliente,
        estado_anterior=estado_anterior,
        estado_nuevo=estado,
        motivo=motivo,
        actualizado_por=request.user
    )

    messages.success(request, "Cliente actualizado exitosamente.")
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
@require_POST
def desasignar_por_cantidad(request):
    cantidad = int(request.POST.get('cantidad', 0))
    usuario_id = request.POST.get('usuario_id')

    if cantidad <= 0 or not usuario_id:
        messages.error(request, "Debe ingresar una cantidad válida y seleccionar un usuario.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    usuario = get_object_or_404(User, id=usuario_id)

    # Obtener clientes pendientes asignados al usuario
    pendientes = Cliente.objects.filter(
        asignado_usuario=usuario,
        estado_actual__id=1  # Estado "pendiente"
    ).order_by('-id')

    total_pendientes = pendientes.count()

    if total_pendientes == 0:
        messages.warning(request, "Este usuario no tiene clientes pendientes asignados.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    if cantidad > total_pendientes:
        messages.warning(request, f"Este usuario solo tiene {total_pendientes} cliente(s) pendientes. Ajuste la cantidad.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    clientes_a_desasignar = pendientes[:cantidad]
    for cliente in clientes_a_desasignar:
        cliente.asignado_usuario = None
        cliente.save()

    messages.success(request, f"{clientes_a_desasignar.count()} clientes desasignados de {usuario.get_full_name()}.")
    return redirect(request.META.get('HTTP_REFERER', '/'))
